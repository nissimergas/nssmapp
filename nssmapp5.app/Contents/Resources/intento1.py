from time import sleep
from bs4 import BeautifulSoup
from selenium import webdriver
import openpyxl
import datetime
from tkinter import *
from requests import get
from json import loads

def get_acciones_bolsa_de_santiago2():
    r=get("http://www.bolsadesantiago.com/mercado/Paginas/cierrebursatil.aspx?RequestAjax=1&hdnPag=1&hdnNombre=&hdnTipo=TODAS")
    lista=loads(r.text)["cierreBursatil"]
    r2=get("http://www.bolsadesantiago.com/mercado/Paginas/cierrebursatil.aspx?RequestAjax=1&hdnPag=2&hdnNombre=&hdnTipo=TODAS")
    lista2=loads(r2.text)["cierreBursatil"]
    r3=get("http://www.bolsadesantiago.com/mercado/Paginas/cierrebursatil.aspx?RequestAjax=1&hdnPag=3&hdnNombre=&hdnTipo=TODAS")
    lista3=loads(r3.text)["cierreBursatil"]
    r4=get("http://www.bolsadesantiago.com/mercado/Paginas/cierrebursatil.aspx?RequestAjax=1&hdnPag=4&hdnNombre=&hdnTipo=TODAS")
    lista4=loads(r4.text)["cierreBursatil"]
    r5=get("http://www.bolsadesantiago.com/mercado/Paginas/cierrebursatil.aspx?RequestAjax=1&hdnPag=5&hdnNombre=&hdnTipo=TODAS")
    lista5=loads(r5.text)["cierreBursatil"]
    r6=get("http://www.bolsadesantiago.com/mercado/Paginas/cierrebursatil.aspx?RequestAjax=1&hdnPag=6&hdnNombre=&hdnTipo=TODAS")
    lista6=loads(r6.text)["cierreBursatil"]
    r7=get("http://www.bolsadesantiago.com/mercado/Paginas/cierrebursatil.aspx?RequestAjax=1&hdnPag=7&hdnNombre=&hdnTipo=TODAS")
    lista7=loads(r7.text)["cierreBursatil"]
    r8=get("http://www.bolsadesantiago.com/mercado/Paginas/cierrebursatil.aspx?RequestAjax=1&hdnPag=8&hdnNombre=&hdnTipo=TODAS")
    lista8=loads(r8.text)["cierreBursatil"]
    r9=get("http://www.bolsadesantiago.com/mercado/Paginas/cierrebursatil.aspx?RequestAjax=1&hdnPag=9&hdnNombre=&hdnTipo=TODAS")
    lista9=loads(r9.text)["cierreBursatil"]
    r10=get("http://www.bolsadesantiago.com/mercado/Paginas/cierrebursatil.aspx?RequestAjax=1&hdnPag=10&hdnNombre=&hdnTipo=TODAS")
    lista10=loads(r10.text)["cierreBursatil"]
    lista.extend(lista2)
    lista.extend(lista3)
    lista.extend(lista4)
    lista.extend(lista5)
    lista.extend(lista6)
    lista.extend(lista7)
    lista.extend(lista8)
    lista.extend(lista9)
    lista.extend(lista10)

    return lista
def string_a_numero(st):
    stn=""
    for letra in st:
        if letra==".":
            pass
        elif letra==",":
            stn+="."
        else:
            stn+=letra
    return float(stn)

def get_information_accion(driver, url,tiempo):
    driver.get(url)
    sleep(tiempo)
    dic={}
    soup=BeautifulSoup(driver.page_source, "html.parser")
    s=soup.find(attrs={"id": "ms-equity-detail-quote"})
    identificador=soup.find(attrs={"class": "symbol"})
    dic["symbol"]=identificador.text
    #print(identificador)
    dic["nombre"]=soup.find("h1",attrs={"class": "name"}).text
    tabla_datos=s.find(attrs={"class": "rtq-grid-bd"})
    last_price=tabla_datos.find(attrs={"domid": "LastPrice"})
    datos=tabla_datos.find_all(attrs={"class": "rtq-grid-row"})

    for dato in datos:
        #print(dato)
        #print(dato["rowid"])
        if dato["rowid"]!="uniquespace":
            #print(dato.find_all(attrs={"class": "rtq-grid-cell-ctn"})[2].text)
            clave=dato["rowid"]
            valor=dato.find_all(attrs={"class": "rtq-grid-cell-ctn"})[2].text
            dic[clave]=valor

    dic["LastPrice"]=last_price.text
    return dic

def get_acciones_bolsa_de_santiago(url):
    datos=[] #dato=(accion, precio, fecha cierre, n neg, volumen, monto, renta, relacion p/u)
    r=get(url)
    soup = BeautifulSoup(r.text, "html.parser")
    cb=soup.find(attrs={"class": "cierreBursatilTablaContenedor"})
    tabla_acciones=cb.find("tbody")
    acciones=tabla_acciones.find_all("tr")
    for accion in acciones:
        caracteristicas=accion.find_all("td")
        nombre=caracteristicas[0].text.strip()
        precio=string_a_numero(caracteristicas[1].text.strip())
        fc=caracteristicas[2].text.strip()
        nneg=string_a_numero(caracteristicas[3].text.strip())
        volumen=string_a_numero(caracteristicas[4].text.strip())
        monto=string_a_numero(caracteristicas[5].text.strip())
        renta=string_a_numero(caracteristicas[6].text.strip())
        rpu=string_a_numero(caracteristicas[7].text.strip())
        dato=(nombre, precio, fc , nneg, volumen, monto, renta, rpu)
        datos.append(dato)
    #print(datos)
    return datos






def retornar_diccionario(url):
     datos={}
     r=get(url)
     start=r.text.find("{")
     #print(start)
     end=len(r.text)-1
     #print(r.text[start:end])
     js=loads(r.text[start:end])
     p_html=js["html"]

     soup = BeautifulSoup(p_html, "html.parser")
     #print(soup.prettify())
     a=soup.find('h1')
     datos["nombre"]=a.text
     msqt_summary=soup.find(attrs={"id": "msqt_summary"})
     header_left=msqt_summary.find(attrs={"id": "header_left"})
     market_wrapper=header_left.find(attrs={"id": "market_wrapper"})
     C_F=header_left.find_all(attrs={"class": "gr_text_bigprice"})
     CR=C_F[0]
     fecha=C_F[1]
     datos["coupon-rate"]=CR.text.strip() #es un porcentaje
     datos["fecha-vencimiento"]=fecha.text.strip()

     tabla=msqt_summary.find_all("table")
     filas=tabla[0].find_all("tr")
     top=filas[0]

     columnas1=top.find_all("td")
     datos["symbol"]=columnas1[0].find("span").text.strip()
     datos["cusip"]=columnas1[1].find("span").text.strip()
     datos["next-call-date"]=columnas1[2].find("span").text.strip()
     #datos["callable"]=columnas1[3].find("span").text.strip()

     #print(tabla[1].find_all("span"))
     columnas2=tabla[1].find_all("span")
     datos["price"]=columnas2[0].text.strip()
     datos["last-trade-yield"]=columnas2[1].text.strip()
     datos["last-trade-date"]=columnas2[2].text.strip()
     datos["us-treasury"]=columnas2[3].text.strip()
     #print(datos)
     return datos


class Gui():
    def createButton(self,width,height,text,location, func):
        self.b=Button(self.frame)
        self.b["text"]=text
        self.b["command"]=func
        self.width=width
        self.height=height
        self.b.pack({'side':location})
        pass
    def __init__(self,root):

        self.root=root

        self.frame=Frame(self.root,width=550,height=250)
        #self.createButton(width=100,height=40,text="Hi",location='left',func=self.pushHi)

        #self.createButton(width=100,height=40,text="bye",location='right',func=self.pushBYE)
        self.lugar=StringVar()
        #self.lugar.set("/Users/nissimergas/Desktop/prueba")
        self.lugar.set("/Users/andresergas/Desktop/mcp/escribe_nombre")
        #self.lugar.set("/Users/nissimergas/desktop/nssmapp3/datos_financieros")
        self.boton=Button(self.frame,text="Acciones",command=self.push_acciones).place(x=10,y=75)
        self.Caja=Entry(self.frame,textvariable=self.lugar).place(x=10,y=40, width=400)
        self.label=Label(self.frame,text="Escriba direccion:").place(x=10,y=10)

        self.boton2=Button(self.frame,text="Bonos",command=self.pushBonos).place(x=10,y=100)
        self.boton3=Button(self.frame,text="Bolsa de Santiago",command=self.push_bolsa_santiago).place(x=10,y=125)
        #my_string_var = tkinter.StringVar(value="Default Value")
        #self.label2=Label(self.frame,textvariable=my_string_var).place(x=10,y=140)
        self.frame.pack()

        #self.label2=Label(self.frame,text="Escriba direccion:").place(x=40,y=70)
        self.root.mainloop()
    def push_acciones(self):
        #phantom="/Users/nissimergas/desktop/nssmapp3/phantomjs"
        phantom="/Users/andresergas/Desktop/mcp/phantomjs"
        driver = webdriver.PhantomJS(executable_path=phantom)
        sleep(1)

        direccion=str(self.lugar.get())+".xlsx"
        if direccion.strip()!=".xlsx":
            excel_document = openpyxl.load_workbook(filename=direccion)
            datos_links = excel_document['Sheet4']
            tabla_precios=excel_document['Sheet3']
            #print(datos_links['A2'].value)
            cantidad_datos=int(datos_links['A2'].value)+1
            now = datetime.datetime.now()
            for n_fila in range(2,cantidad_datos+1):
                ejecutar=int(datos_links.cell(row=n_fila, column=4).value)
                if ejecutar==1:
                    try:
                        url=datos_links.cell(row=n_fila, column=3).value
                        tiempo=int(datos_links['A4'].value)
                        dic=get_information_accion(driver, url,tiempo)
                        symbol=dic["symbol"]
                        nombre=dic["nombre"]
                        last_price=dic["LastPrice"]
                        if last_price!='--':
                            try:
                                last_price=float(string_a_numero(last_price))
                            except:
                                "not number"
                            #last_price=float(last_price)

                        week52H=dic["st168"]
                        if week52H!='--':
                            try:
                                week52H=float(string_a_numero(week52H))
                            except:
                                "not number"
                            #week52H=float(week52H)
                        week52L=dic["st169"]
                        if week52L!='--':
                            try:
                                week52L=float(string_a_numero(week52L))
                            except:
                                "not number"

                        week52H_day=dic["st109"]
                        week52L_day=dic["st106"]
                        prev_close=dic['ClosePrice']
                        if prev_close!='--':
                            try:
                                prev_close=float(string_a_numero(prev_close))
                            except:
                                "not number"

                        symbolos=symbol.split(":")
                        tabla_precios.cell(row=n_fila, column=1).value = symbolos[0]
                        tabla_precios.cell(row=n_fila, column=2).value = symbolos[1]
                        tabla_precios.cell(row=n_fila, column=3).value = nombre
                        tabla_precios.cell(row=n_fila, column=4).value = last_price
                        tabla_precios.cell(row=n_fila, column=5).value = prev_close
                        tabla_precios.cell(row=n_fila, column=6).value = week52H_day
                        tabla_precios.cell(row=n_fila, column=7).value = week52H
                        tabla_precios.cell(row=n_fila, column=8).value = week52L_day
                        tabla_precios.cell(row=n_fila, column=9).value = week52L
                        tabla_precios.cell(row=n_fila, column=10).value = now
                    except:
                        tabla_precios.cell(row=n_fila, column=1).value ="error"
                        tabla_precios.cell(row=n_fila, column=2).value ="error"
                        tabla_precios.cell(row=n_fila, column=3).value ="error"
                        tabla_precios.cell(row=n_fila, column=4).value ="error"
                        tabla_precios.cell(row=n_fila, column=5).value ="error"
                        tabla_precios.cell(row=n_fila, column=6).value ="error"
                        tabla_precios.cell(row=n_fila, column=7).value ="error"
                        tabla_precios.cell(row=n_fila, column=8).value ="error"
                        tabla_precios.cell(row=n_fila, column=9).value ="error"
                        tabla_precios.cell(row=n_fila, column=10).value ="error"

            excel_document.save(direccion)

    def pushBonos(self):
        #print(self.lugar.get())
        #cwd = os.getcwd()
        direccion=str(self.lugar.get())+".xlsx"
        if direccion.strip()!=".xlsx":
            excel_document = openpyxl.load_workbook(filename=direccion)
            datos_links = excel_document['Sheet2']
            tabla_precios=excel_document['Sheet1']

            cantidad_datos=int(datos_links['A2'].value)+1
            for n_fila in range(2,cantidad_datos+1):
                url=datos_links.cell(row=n_fila, column=3).value
                if url!=None:
                    #print("_______________________________")
                    #print(url)
                    try:
                        dic=retornar_diccionario(url)
                        tabla_precios.cell(row=n_fila, column=2).value = dic["cusip"]
                        tabla_precios.cell(row=n_fila, column=3).value = dic["symbol"]
                        tabla_precios.cell(row=n_fila, column=4).value = dic["nombre"]
                        tabla_precios.cell(row=n_fila, column=5).value = dic["fecha-vencimiento"]
                        CR=dic["coupon-rate"]
                        if "—" not in CR:
                            try:
                                CR=float(CR)
                            except:
                                print(CR)
                        tabla_precios.cell(row=n_fila, column=6).value = CR
                        tabla_precios.cell(row=n_fila, column=7).value = dic["last-trade-date"]
                        precio=dic["price"][1:]
                        #print(precio)
                        if "—" not in precio:
                            try:
                                precio=float(precio)
                            except:
                                print(precio)

                        tabla_precios.cell(row=n_fila, column=8).value = precio
                        L_T_Y=dic["last-trade-yield"].strip("%")
                        #print(L_T_Y)
                        if "—" not in L_T_Y :
                            try:
                                L_T_Y=float(L_T_Y)
                            except:
                                print(L_T_Y)

                        tabla_precios.cell(row=n_fila, column=9).value = L_T_Y
                    except:
                        tabla_precios.cell(row=n_fila, column=2).value ="error"
                        tabla_precios.cell(row=n_fila, column=3).value ="error"
                        tabla_precios.cell(row=n_fila, column=4).value ="error"
                        tabla_precios.cell(row=n_fila, column=5).value ="error"
                        tabla_precios.cell(row=n_fila, column=6).value ="error"
                        tabla_precios.cell(row=n_fila, column=7).value ="error"
                        tabla_precios.cell(row=n_fila, column=8).value ="error"
                        tabla_precios.cell(row=n_fila, column=9).value ="error"


            excel_document.save(direccion)
    def push_bolsa_santiago(self):
        direccion=str(self.lugar.get())+".xlsx"
        if direccion.strip()!=".xlsx":
            excel_document = openpyxl.load_workbook(filename=direccion)
            tabla_precios=excel_document['bolsa_santiago']
            datos=get_acciones_bolsa_de_santiago2()

            n_fila=2
            for fila in datos:
                tabla_precios.cell(row=n_fila, column=1).value = fila["Nemo"]
                tabla_precios.cell(row=n_fila, column=2).value = fila["PrecioCierre"]
                tabla_precios.cell(row=n_fila, column=3).value = fila["FechaCierreString"]
                #tabla_precios.cell(row=n_fila, column=4).value = fila[3]
                #tabla_precios.cell(row=n_fila, column=5).value = fila[4]
                #tabla_precios.cell(row=n_fila, column=6).value = fila[5]
                #tabla_precios.cell(row=n_fila, column=7).value = fila[6]
                #tabla_precios.cell(row=n_fila, column=8).value = fila[7]
                n_fila+=1


            excel_document.save(direccion)
    def push_bolsa_santiago2(self):
        direccion=str(self.lugar.get())+".xlsx"
        if direccion.strip()!=".xlsx":
            excel_document = openpyxl.load_workbook(filename=direccion)
            tabla_precios=excel_document['bolsa_santiago']
            datos=get_acciones_bolsa_de_santiago("http://www.bolsadesantiago.com/mercado/Paginas/cierrebursatil.aspx")

            n_fila=2
            for fila in datos:
                tabla_precios.cell(row=n_fila, column=1).value = fila[0]
                tabla_precios.cell(row=n_fila, column=2).value = fila[1]
                tabla_precios.cell(row=n_fila, column=3).value = fila[2]
                tabla_precios.cell(row=n_fila, column=4).value = fila[3]
                tabla_precios.cell(row=n_fila, column=5).value = fila[4]
                tabla_precios.cell(row=n_fila, column=6).value = fila[5]
                tabla_precios.cell(row=n_fila, column=7).value = fila[6]
                tabla_precios.cell(row=n_fila, column=8).value = fila[7]
                n_fila+=1


            excel_document.save(direccion)



if __name__== '__main__':
     #get_acciones_bolsa_de_santiago("http://www.bolsadesantiago.com/mercado/Paginas/cierrebursatil.aspx")
     #get_acciones_bolsa_de_santiago2()
     root=Tk()
     root.title("nssmapp4")
     root.geometry("500x400")
     Gui(root)